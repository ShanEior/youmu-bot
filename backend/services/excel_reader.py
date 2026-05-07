from pathlib import Path
import zipfile
from xml.etree import ElementTree

from openpyxl import load_workbook

from .merged_cell_handler import fill_merged_cells

CUSTOM_PROPERTIES_PATH = "docProps/custom.xml"
VT_NAMESPACE = "{http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes}"
PROPERTY_NAMESPACE = "{http://schemas.openxmlformats.org/officeDocument/2006/custom-properties}"


class ExcelReadError(Exception):
    pass


def read_workbook(file_path):
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xls":
        raise ExcelReadError("当前 MVP 暂不支持 .xls，请转换为 .xlsx 后上传")

    if suffix != ".xlsx":
        raise ExcelReadError(f"不支持的 Excel 文件格式: {suffix or 'unknown'}")

    sanitized_path = sanitize_workbook_for_openpyxl(path)
    workbook = load_workbook(sanitized_path, data_only=True, read_only=False)
    try:
        sheets = []

        for sheet_index, worksheet in enumerate(workbook.worksheets):
            if sheet_index == 0:
                continue

            data = read_sheet_data(worksheet)
            sheets.append({
                "sheet_name": worksheet.title,
                "data": data,
            })

        return sheets
    finally:
        workbook.close()
        if sanitized_path != path and sanitized_path.exists():
            sanitized_path.unlink()


def sanitize_workbook_for_openpyxl(file_path):
    with zipfile.ZipFile(file_path, "r") as source_zip:
        try:
            custom_xml = source_zip.read(CUSTOM_PROPERTIES_PATH)
        except KeyError:
            return file_path

        sanitized_xml = sanitize_custom_properties_xml(custom_xml)
        if sanitized_xml == custom_xml:
            return file_path

        temp_path = file_path.with_name(f"{file_path.stem}_sanitized{file_path.suffix}")
        with zipfile.ZipFile(temp_path, "w") as target_zip:
            for item in source_zip.infolist():
                data = sanitized_xml if item.filename == CUSTOM_PROPERTIES_PATH else source_zip.read(item.filename)
                target_zip.writestr(item, data)

        return temp_path


def sanitize_custom_properties_xml(xml_bytes):
    root = ElementTree.fromstring(xml_bytes)
    removed = False

    for property_node in list(root.findall(f"{PROPERTY_NAMESPACE}property")):
        name = property_node.get("name")
        if name is None:
            root.remove(property_node)
            removed = True
            continue

        value_nodes = list(property_node)
        if not value_nodes:
            root.remove(property_node)
            removed = True
            continue

        value_node = value_nodes[0]
        if not value_node.tag.startswith(VT_NAMESPACE):
            root.remove(property_node)
            removed = True

    if not removed:
        return xml_bytes

    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def read_sheet_data(worksheet):
    filled_values = fill_merged_cells(worksheet)
    rows = []

    for row in worksheet.iter_rows():
        row_values = []
        for cell in row:
            value = filled_values.get((cell.row, cell.column), cell.value)
            row_values.append(format_cell_value(value))
        rows.append(trim_empty_tail(row_values))

    return trim_empty_rows(rows)


def format_cell_value(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def trim_empty_tail(row):
    values = list(row)
    while values and values[-1] == "":
        values.pop()
    return values


def trim_empty_rows(rows):
    values = list(rows)
    while values and not any(values[-1]):
        values.pop()
    return values
