from openpyxl.utils import get_column_letter

DIRECTION_KEYWORDS = {
    "start": ("起点", "采集", "start", "输入"),
    "end": ("终点", "产品", "end", "输出"),
}

FIELD_KEYWORDS = {
    "connector": ("插件", "连接器", "connector"),
    "pin": ("针脚", "针位", "脚位", "孔位", "针", "引脚", "位", "pin"),
    "content": ("内容", "信号", "说明"),
}

SIGNAL_KEYWORDS = ("信号性质", "信号规格", "线径", "规格")
REMARK_KEYWORDS = ("备注", "说明", "特殊")
REQUIRED_FIELDS = ("start_connector", "start_pin", "end_connector", "end_pin")
ALL_FIELDS = (
    "start_connector",
    "start_pin",
    "end_connector",
    "end_pin",
    "start_content",
    "end_content",
    "signal_type",
    "remark",
)
FIELD_ALIAS_BY_CONNECTOR = {
    "start_connector": "start_pin",
    "end_connector": "end_pin",
}
HEADER_SCAN_LIMIT = 6


def detect_columns(rows):
    best = None

    for index, row in enumerate(rows):
        best = choose_better_candidate(best, build_candidate(index + 1, detect_row_columns(row)))

    if not best or best["required_score"] < len(REQUIRED_FIELDS):
        header_limit = min(len(rows), HEADER_SCAN_LIMIT)
        for index in range(header_limit - 1):
            combined_row = combine_header_rows(rows[index], rows[index + 1])
            best = choose_better_candidate(best, build_candidate(index + 2, detect_row_columns(combined_row)))

    if not best:
        return {
            "valid": False,
            "header_row": None,
            "matched_columns": {},
        }

    return {
        "valid": all(field in best["matched_columns"] for field in REQUIRED_FIELDS),
        "header_row": best["header_row"],
        "matched_columns": best["matched_columns"],
    }


def build_candidate(header_row, matched):
    required_score = sum(1 for field in REQUIRED_FIELDS if field in matched)
    if required_score == 0:
        return None

    return {
        "header_row": header_row,
        "matched_columns": matched,
        "score": sum(1 for field in ALL_FIELDS if field in matched),
        "required_score": required_score,
    }


def choose_better_candidate(best, candidate):
    if not candidate:
        return best

    if not best or (candidate["required_score"], candidate["score"], candidate["header_row"]) > (
        best["required_score"],
        best["score"],
        best["header_row"],
    ):
        return candidate

    return best


def combine_header_rows(upper_row, lower_row):
    max_length = max(len(upper_row), len(lower_row))
    combined = []

    for column_index in range(max_length):
        upper_text = get_normalized_value(upper_row, column_index)
        lower_text = get_normalized_value(lower_row, column_index)
        combined.append(f"{upper_text}{lower_text}")

    return combined


def get_normalized_value(row, column_index):
    if column_index >= len(row):
        return ""
    return normalize_text(row[column_index])


def detect_row_columns(row):
    matched = {}

    for column_index, value in enumerate(row):
        text = normalize_text(value)
        if not text:
            continue

        side = detect_side(text)
        field_type = detect_field_type(text)

        if side and field_type:
            field = f"{side}_{field_type}"
            matched.setdefault(field, get_column_letter(column_index + 1))
            continue

        if any(keyword in text for keyword in SIGNAL_KEYWORDS):
            matched.setdefault("signal_type", get_column_letter(column_index + 1))
            continue

        if any(keyword in text for keyword in REMARK_KEYWORDS):
            matched.setdefault("remark", get_column_letter(column_index + 1))

    fill_missing_pin_columns(matched)
    return matched


def fill_missing_pin_columns(matched):
    for connector_field, pin_field in FIELD_ALIAS_BY_CONNECTOR.items():
        if connector_field not in matched or pin_field in matched:
            continue

        connector_column = matched[connector_field]
        pin_column = get_column_letter(column_index_from_letter(connector_column) + 1)
        matched[pin_field] = pin_column


def column_index_from_letter(column_letter):
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def detect_side(text):
    for side, keywords in DIRECTION_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return side
    return None


def detect_field_type(text):
    for field_type, keywords in FIELD_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return field_type
    return None


def normalize_text(value):
    return str(value or "").strip().replace(" ", "").replace("\n", "").lower()
