from openpyxl import Workbook


HEADERS = ["Net", "Sub", "Start", "End", "Remark"]
COLUMN_WIDTHS = {
    "A": 12,
    "B": 12,
    "C": 24,
    "D": 24,
    "E": 20,
}


def write_converted_workbook(output_path, sheets_data):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_data in sheets_data:
        worksheet = workbook.create_sheet(title=sheet_data["sheet_name"])
        worksheet.append(HEADERS)

        for row in sheet_data.get("rows", []):
            worksheet.append([
                row.get("net", ""),
                row.get("sub", ""),
                row.get("start", ""),
                row.get("end", ""),
                row.get("remark", ""),
            ])

        for column, width in COLUMN_WIDTHS.items():
            worksheet.column_dimensions[column].width = width

    if not workbook.sheetnames:
        worksheet = workbook.create_sheet(title="Sheet1")
        worksheet.append(HEADERS)
        for column, width in COLUMN_WIDTHS.items():
            worksheet.column_dimensions[column].width = width

    workbook.save(output_path)
